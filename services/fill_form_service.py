import os
import re
import zipfile
import tempfile
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

# ==============================================================================
# HÀM HỖ TRỢ XỬ LÝ CHUỖI & DÒNG EXCEL
# ==============================================================================
def clean_filename(filename):
    """Loại bỏ ký tự không hợp lệ trong tên file Windows"""
    return re.sub(r'[\\/*?:"<>|]', '_', str(filename)).strip()

def clean_slash(value):
    """Nếu giá trị là '/' (hoặc có chứa khoảng trắng xung quanh '/') thì đổi thành 'NA'"""
    if pd.isna(value) or value is None:
        return ""
    val_str = str(value).strip()
    # Kiểm tra nếu sau khi xóa khoảng trắng chỉ còn đúng 1 ký tự '/'
    if val_str == "/":
        return "NA"
    return val_str

def format_date_mmm(value):
    """Chuyển đổi ngày tháng về định dạng DD/MMM/YYYY (ví dụ: 15/Jan/2026)"""
    if pd.isna(value) or value is None or str(value).strip() == "/":
        return ""
    dt = pd.to_datetime(value, errors='coerce')
    if pd.isna(dt):
        return str(value).strip()
    return dt.strftime("%d/%b/%Y")

def set_wrap_text(cell):
    """Bật tính năng xuống dòng tự động (Wrap Text) giữ nguyên căn lề cũ"""
    current_alignment = cell.alignment
    cell.alignment = Alignment(
        horizontal=current_alignment.horizontal,
        vertical=current_alignment.vertical,
        wrap_text=True
    )

def adjust_row_height(ws, row_idx, texts, char_limit_per_line=25, base_height=20, line_height=15):
    """Tự động tính toán và chỉnh chiều cao dòng chuẩn xác khi chữ dài HOẶC có chứa dấu \n"""
    max_lines = 1
    for txt in texts:
        if txt:
            paragraphs = str(txt).split('\n')
            total_lines_for_cell = 0
            for p in paragraphs:
                p_lines = (len(p) // char_limit_per_line) + 1
                total_lines_for_cell += p_lines
            if total_lines_for_cell > max_lines:
                max_lines = total_lines_for_cell
    
    if max_lines > 1:
        new_height = base_height + (max_lines - 1) * line_height
        current_h = ws.row_dimensions[row_idx].height or base_height
        ws.row_dimensions[row_idx].height = max(current_h, new_height)

def get_row_index(cell_name):
    """Lấy chỉ số dòng từ tên ô (ví dụ 'J11' -> 11)"""
    match = re.search(r'\d+', cell_name)
    return int(match.group()) if match else 1

# ==============================================================================
# QUÉT DỮ LIỆU TỪ TỆP PDF
# ==============================================================================
def extract_code_from_pdf(pdf_file_obj):
    """Đọc PDF từ bộ nhớ/máy tạm và nhặt tất cả mã quy trình tại Mục 4"""
    try:
        with pdfplumber.open(pdf_file_obj) as pdf:
            page = pdf.pages[1] if len(pdf.pages) > 1 else pdf.pages[0]
            text = page.extract_text()
            if not text:
                return ""

            lines = text.split('\n')
            found_section_4 = False
            extracted_codes = []

            for line in lines:
                line_str = line.strip()
                if "4. Tài liệu cơ sở" in line_str or "Reference documents" in line_str:
                    found_section_4 = True
                    continue

                if found_section_4 and (line_str.startswith("5.") or "5. Thiết bị" in line_str or "5. Local" in line_str):
                    break

                if found_section_4 and line_str:
                    match = re.search(r'([A-Za-z0-9]+/[A-Za-z0-9\(\)\-]+)', line_str)
                    if match:
                        code = match.group(1)
                        if code not in extracted_codes:
                            extracted_codes.append(code)

            return "\n".join(extracted_codes)
    except Exception as e:
        return ""

def scan_pdf_files(pdf_files):
    """Tạo Dictionary map: {Mã_GCN: Mã_Tài_Liệu} từ danh sách tệp PDF"""
    pdf_map = {}
    for file in pdf_files:
        ma_gcn = os.path.splitext(file.name)[0].strip()
        ma_tai_lieu = extract_code_from_pdf(file)
        pdf_map[ma_gcn] = ma_tai_lieu
    return pdf_map

# ==============================================================================
# HÀM BỎ BIẾN ĐỔI CHUỖI THEO LOẠI (TRANSFORMATION ENGINE)
# ==============================================================================
def apply_transformation(val, transform_type, raw_id="", pdf_data_map=None):
    """
    Áp dụng logic biến đổi dữ liệu tùy theo Kiểu (Transform Type) được chọn trên UI
    """
    # 1. Làm sạch giá trị ban đầu (chuyển '/' thành 'NA')
    str_val = clean_slash(val)

    # 2. Xử lý các logic đặc biệt
    if transform_type == "Cắt lấy phần sau dấu '/'":
        if str_val == "NA":
            return "NA"
        return str_val.split("/")[-1].strip() if "/" in str_val else str_val

    elif transform_type == "Tạo mã 'M' (Prefix M)":
        parts = raw_id.split('.')
        return f"M{parts[1]}" if len(parts) >= 2 else "M"

    elif transform_type == "Định dạng Ngày (DD/MMM/YYYY)":
        return format_date_mmm(val)

    elif transform_type == "Tra cứu PDF theo Mã GCN":
        return pdf_data_map.get(str_val, "") if pdf_data_map else ""

    elif transform_type == "Viết HOA toàn bộ":
        return str_val.upper()

    elif transform_type == "Viết thường toàn bộ":
        return str_val.lower()

    elif transform_type == "Đánh tích nhóm '6' (ü)":
        parts = str(raw_id).split('.')
        if len(parts) >= 3 and parts[2].strip().startswith('6'):
            return "ü"
        return ""

    # Mặc định (Nguyên bản Direct): Trả về giá trị đã làm sạch ('/' -> 'NA')
    return str_val

# ==============================================================================
# ENGINE XỬ LÝ CHÍNH
# ==============================================================================
def run_generate_forms(file_tong_bytes, file_form_bytes, pdf_files, mapping_config):
    """
    Xử lý tạo các file Form Excel và đóng gói vào file ZIP (Kèm file Log đối chiếu PDF)
    """
    try:
        # Danh sách lưu các dòng nhật ký (Log)
        log_lines = []
        log_lines.append(f"=== NHẬT KÝ ĐỐI CHIẾU DỮ LIỆU PDF & EXCEL ===")
        log_lines.append(f"Thời gian thực hiện: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        # 1. Quét PDF trước nếu có
        pdf_data_map = {}
        if pdf_files:
            pdf_data_map = scan_pdf_files(pdf_files)
            log_lines.append(f"-> Đã quét tổng cộng {len(pdf_files)} tệp PDF.")
        else:
            log_lines.append("-> CẢNH BÁO: Không có tệp PDF nào được tải lên.\n")

        # 2. Đọc file Tổng
        df = pd.read_excel(file_tong_bytes)
        df = df.replace(r'^\s*/\s*$', 'NA', regex=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = tempfile.mkdtemp()
        generated_files = []

        log_lines.append("\n=== CHI TIẾT TRẠNG THÁI TỪNG FILE ===")

        # 3. Lặp từng dòng dữ liệu trong File Tổng
        for idx, row in df.iterrows():
            id_col = mapping_config.get("id_column")
            if not id_col or id_col not in row or pd.isna(row[id_col]):
                continue
            
            raw_id = str(row[id_col]).strip()
            if not raw_id or raw_id.lower() == 'nan':
                continue

            # Mở workbook form mẫu
            file_form_bytes.seek(0)
            wb = load_workbook(file_form_bytes)
            ws = wb.active

            # Lưu log vết kiểm tra PDF cho dòng này
            pdf_status_info = []

            # --- A. Xử lý danh sách Dynamic Mapping Pairs ---
            for pair in mapping_config.get("dynamic_pairs", []):
                col_name = pair.get("excel_col")
                target_cells_str = pair.get("target_cell", "")
                target_cells = [c.strip().upper() for c in target_cells_str.split(",") if c.strip()]
                transform_type = pair.get("transform_type", "Nguyên bản (Direct)")

                if col_name in row and target_cells:
                    raw_val = row[col_name]
                    final_val = apply_transformation(raw_val, transform_type, raw_id=raw_id, pdf_data_map=pdf_data_map)

                    # Ghi nhận log nếu đây là cột tra cứu PDF
                    if transform_type == "Tra cứu PDF theo Mã GCN":
                        ma_gcn = clean_slash(raw_val)
                        if not ma_gcn or ma_gcn == "NA":
                            pdf_status_info.append("Không có Mã GCN")
                        elif ma_gcn in pdf_data_map:
                            if pdf_data_map[ma_gcn]:
                                pdf_status_info.append(f" [Thành công] Tìm thấy PDF cho GCN '{ma_gcn}'")
                            else:
                                pdf_status_info.append(f" [Cảnh báo] Có file PDF '{ma_gcn}' nhưng không trích xuất được mã Mục 4")
                        else:
                            pdf_status_info.append(f"❌ [THIẾU PDF] Không tìm thấy file PDF tương ứng với Mã GCN '{ma_gcn}'")

                    for cell_name in target_cells:
                        ws[cell_name] = final_val
                        set_wrap_text(ws[cell_name])
                        adjust_row_height(ws, row_idx=get_row_index(cell_name), texts=[final_val])

            # Ghi dòng nhật ký cho Mã QL hiện tại
            log_detail = f"- Mã QL: {raw_id}"
            if pdf_status_info:
                log_detail += " | " + " ; ".join(pdf_status_info)
            else:
                log_detail += " | Tạo file thành công"
            log_lines.append(log_detail)

            # Lưu file Excel theo Mã QL
            safe_filename = clean_filename(raw_id)
            out_file_path = os.path.join(temp_dir, f"{safe_filename}.xlsx")
            wb.save(out_file_path)
            wb.close()
            generated_files.append(out_file_path)

        if not generated_files:
            return None, "Không tạo được file nào. Vui lòng kiểm tra lại cột định danh Mã quản lý."

        # Tạo file Log (.txt) trong thư mục tạm
        log_file_path = os.path.join(temp_dir, "Log_Doi_Chieu_PDF.txt")
        with open(log_file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(log_lines))

        # 4. Nén tất cả các file Excel + File Log vào ZIP
        zip_path = os.path.join(tempfile.gettempdir(), f"Danh_Sach_Form_Hoan_Thanh_{timestamp}.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Thêm các file Excel
            for file_path in generated_files:
                zipf.write(file_path, arcname=os.path.basename(file_path))
            # Thêm file Log vào file ZIP
            zipf.write(log_file_path, arcname="Log_Doi_Chieu_PDF.txt")

        # Dọn dẹp temp
        for f in generated_files + [log_file_path]:
            try: os.remove(f)
            except Exception: pass

        return zip_path, f"Đã xuất thành công {len(generated_files)} file Form Excel và 1 file Log!"

    except Exception as e:
        return None, f"Lỗi xử lý: {str(e)}"