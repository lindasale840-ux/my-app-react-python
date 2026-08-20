import os
import re
import json
import zipfile
import tempfile
import io
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse

router = APIRouter(prefix="/api/excel", tags=["ExcelProcess"])

# ==============================================================================
# HÀM HỖ TRỢ XỬ LÝ CHUỖI & DÒNG EXCEL
# ==============================================================================
def clean_filename(filename):
    """Loại bỏ ký tự không hợp lệ trong tên file Windows"""
    return re.sub(r'[\\/*?:"<>|]', '_', str(filename)).strip()

def clean_slash(value):
    if pd.isna(value) or value is None:
        return ""
    val_str = str(value).strip()
    if val_str == "/" or val_str == "":
        return "NA" if val_str == "/" else ""
    return val_str

def format_date_mmm(value):
    """Chuyển đổi ngày tháng về định dạng DD/MMM/YYYY (ví dụ: 15/Jan/2026)"""
    if pd.isna(value) or value is None or str(value).strip() == "/":
        return ""
    dt = pd.to_datetime(value, errors='coerce')
    if pd.isna(dt):
        return str(value).strip()
    return dt.strftime("%d/%b/%Y")

def set_wrap_text(cell):
    """Bật tính năng xuống dòng tự động (Wrap Text) giữ nguyên căn lề cũ"""
    current_alignment = cell.alignment
    cell.alignment = Alignment(
        horizontal=current_alignment.horizontal,
        vertical=current_alignment.vertical,
        wrap_text=True
    )

def adjust_row_height(ws, row_idx, texts, char_limit_per_line=25, base_height=20, line_height=15):
    """Tự động tính toán và chỉnh chiều cao dòng chuẩn xác"""
    max_lines = 1
    for txt in texts:
        if txt:
            paragraphs = str(txt).split('\n')
            total_lines_for_cell = 0
            for p in paragraphs:
                p_lines = (len(p) // char_limit_per_line) + 1
                total_lines_for_cell += p_lines
            if total_lines_for_cell > max_lines:
                max_lines = total_lines_for_cell
    
    if max_lines > 1:
        new_height = base_height + (max_lines - 1) * line_height
        current_h = ws.row_dimensions[row_idx].height or base_height
        ws.row_dimensions[row_idx].height = max(current_h, new_height)

def get_row_index(cell_name):
    """Lấy chỉ số dòng từ tên ô (ví dụ 'J11' -> 11)"""
    match = re.search(r'\d+', cell_name)
    return int(match.group()) if match else 1

# ==============================================================================
# QUÉT DỮ LIỆU TỪ TỆP PDF
# ==============================================================================
def extract_code_from_pdf(pdf_file_obj):
    """Đọc PDF từ bộ nhớ/máy tạm và nhặt tất cả mã quy trình tại Mục 4"""
    try:
        with pdfplumber.open(pdf_file_obj) as pdf:
            page = pdf.pages[1] if len(pdf.pages) > 1 else pdf.pages[0]
            text = page.extract_text()
            if not text:
                return ""

            lines = text.split('\n')
            found_section_4 = False
            extracted_codes = []

            for line in lines:
                line_str = line.strip()
                if "4. Tài liệu cơ sở" in line_str or "Reference documents" in line_str:
                    found_section_4 = True
                    continue

                if found_section_4 and (line_str.startswith("5.") or "5. Thiết bị" in line_str or "5. Local" in line_str):
                    break

                if found_section_4 and line_str:
                    match = re.search(r'\b([A-Za-z0-9]+(?:[/\-_][A-Za-z0-9\(\)\-]+)+)\b', line_str)
                    if match:
                        code = match.group(1)
                        if code not in extracted_codes:
                            extracted_codes.append(code)

            return "\n".join(extracted_codes)
    except Exception:
        return ""

def scan_pdf_files(pdf_files):
    """Tạo Dictionary map: {Mã_GCN: Mã_Tài_Liệu} và danh sách Log quét PDF"""
    pdf_map = {}
    pdf_logs = []
    
    total_files = len(pdf_files)
    pdf_logs.append(f"--- ĐANG QUÉT {total_files} FILE PDF TRONG THƯ MỤC ---")
    
    for idx, file in enumerate(pdf_files, 1):
        ma_gcn = os.path.splitext(file.filename)[0].strip()
        ma_tai_lieu = extract_code_from_pdf(file.file)
        pdf_map[ma_gcn] = ma_tai_lieu
        
        ma_tl_display = ma_tai_lieu.replace('\n', ' | ') if ma_tai_lieu else ""
        pdf_logs.append(f"[{idx}/{total_files}] PDF: {ma_gcn} -> Mã TL (I20): [{ma_tl_display}]")
        
    return pdf_map, pdf_logs

# ==============================================================================
# ENGINE BIẾN ĐỔI CHUỖI
# ==============================================================================
def apply_transformation(val, transform_type, raw_id="", pdf_data_map=None):
    str_val = clean_slash(val)

    if transform_type == "Cắt lấy phần sau dấu '/'":
        if str_val == "NA":
            return "NA"
        return str_val.split("/")[-1].strip() if "/" in str_val else str_val

    elif transform_type == "Tạo mã 'M' (Prefix M)":
        parts = raw_id.split('.')
        return f"M{parts[1]}" if len(parts) >= 2 else "M"

    elif transform_type == "Định dạng Ngày (DD/MMM/YYYY)":
        return format_date_mmm(val)

    elif transform_type == "Tra cứu PDF theo Mã GCN":
        return pdf_data_map.get(str_val, "") if pdf_data_map else ""

    elif transform_type == "Viết HOA toàn bộ":
        return str_val.upper()

    elif transform_type == "Viết thường toàn bộ":
        return str_val.lower()

    elif transform_type == "Đánh tích nhóm '6' (ü)":
        parts = str(raw_id).split('.')
        if len(parts) >= 3 and parts[2].strip().startswith('6'):
            return "ü"
        return ""

    return str_val

# ==============================================================================
# API ENDPOINTS
# ==============================================================================
@router.post("/preview-columns")
async def preview_columns(file_tong: UploadFile = File(...)):
    """API Đọc danh sách các cột từ File Tổng Excel để trả về cho Frontend"""
    try:
        contents = await file_tong.read()
        df_preview = pd.read_excel(io.BytesIO(contents), nrows=1)
        columns = [str(col) for col in df_preview.columns]
        return JSONResponse(content={"columns": columns})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không thể đọc file Tổng: {str(e)}")

@router.post("/process-form")
async def process_excel_form(
    file_tong: UploadFile = File(...),
    file_form: UploadFile = File(...),
    pdf_files: List[UploadFile] = File(default=[]),
    config: str = Form(...)
):
    """API chính để điền form Excel, trích xuất PDF và đóng gói ZIP"""
    try:
        mapping_config = json.loads(config)
        log_lines = []

        # 1. Quét PDF nếu có
        pdf_data_map = {}
        if pdf_files:
            pdf_data_map, pdf_scan_logs = scan_pdf_files(pdf_files)
            log_lines.extend(pdf_scan_logs)
            log_lines.append("")
        else:
            log_lines.append("--- KHÔNG CÓ FILE PDF NÀO ĐƯỢC TẢI LÊN ---\n")

        # 2. Đọc File Tổng
        file_tong_bytes = io.BytesIO(await file_tong.read())
        df = pd.read_excel(file_tong_bytes)
        df = df.replace(r'^\s*/\s*$', 'NA', regex=True)

        id_col = mapping_config.get("id_column")
        valid_rows = [r for _, r in df.iterrows() if id_col in r and pd.notna(r[id_col]) and str(r[id_col]).strip() and str(r[id_col]).strip().lower() != 'nan']
        total_excel_files = len(valid_rows)

        log_lines.append("--- ĐANG ĐỌC FILE TỔNG EXCEL VÀ ĐIỀN DỮ LIỆU ---")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = tempfile.mkdtemp()
        generated_files = []

        form_bytes_content = await file_form.read()

        # 3. Vòng lặp dữ liệu
        file_count = 0
        for idx, row in df.iterrows():
            if id_col not in row or pd.isna(row[id_col]):
                continue
            
            raw_id = str(row[id_col]).strip()
            if not raw_id or raw_id.lower() == 'nan':
                continue

            file_count += 1

            # Mở template form
            wb = load_workbook(io.BytesIO(form_bytes_content))
            ws = wb.active
            i20_val_log = ""

            # Mapping động
            for pair in mapping_config.get("dynamic_pairs", []):
                col_name = pair.get("excel_col")
                target_cells_str = pair.get("target_cell", "")
                target_cells = [c.strip().upper() for c in target_cells_str.split(",") if c.strip()]
                transform_type = pair.get("transform_type", "Nguyên bản (Direct)")

                if col_name in row and target_cells:
                    raw_val = row[col_name]
                    final_val = apply_transformation(raw_val, transform_type, raw_id=raw_id, pdf_data_map=pdf_data_map)

                    if transform_type == "Tra cứu PDF theo Mã GCN" or "I20" in target_cells:
                        i20_val_log = final_val.replace('\n', ' | ')

                    for cell_name in target_cells:
                        ws[cell_name] = final_val
                        set_wrap_text(ws[cell_name])
                        adjust_row_height(ws, row_idx=get_row_index(cell_name), texts=[final_val])

            # Checkmark special rules
            rules = mapping_config.get("special_rules", {})
            if rules.get("checkmark_logic", {}).get("active"):
                cell_k = rules["checkmark_logic"].get("cell_option1", "K14")
                cell_n = rules["checkmark_logic"].get("cell_option2", "N14")
                
                parts = raw_id.split('.')
                ws[cell_k] = ""
                ws[cell_n] = ""
                
                if len(parts) >= 3 and parts[2].strip().startswith('6'):
                    ws[cell_k] = "ü"
                else:
                    ws[cell_n] = "ü"

            # Lưu file
            safe_filename = clean_filename(raw_id)
            out_file_path = os.path.join(temp_dir, f"{safe_filename}.xlsx")
            wb.save(out_file_path)
            wb.close()
            generated_files.append(out_file_path)

            log_lines.append(f"[{file_count}/{total_excel_files}] Hoàn tất: {safe_filename}.xlsx (I20: '{i20_val_log}')")

        if not generated_files:
            raise HTTPException(status_code=400, detail="Không tạo được file nào. Kiểm tra cột Mã Quản Lý.")

        # Tạo file Log txt
        log_file_path = os.path.join(temp_dir, "Log_Doi_Chieu_PDF.txt")
        with open(log_file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(log_lines))

        # Đóng gói ZIP
        zip_filename = f"Form_Excel_Completed_{timestamp}.zip"
        zip_path = os.path.join(tempfile.gettempdir(), zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in generated_files:
                zipf.write(file_path, arcname=os.path.basename(file_path))
            zipf.write(log_file_path, arcname="Log_Doi_Chieu_PDF.txt")

        return FileResponse(
            path=zip_path,
            filename=zip_filename,
            media_type="application/zip"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý hệ thống: {str(e)}")